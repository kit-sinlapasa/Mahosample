import csv
from datetime import datetime
from io import BytesIO, StringIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.import_job import ImportJob, ImportJobRow
from app.models.sample_request import RequestStatus, ShippingStatus
from app.services.sample_request_service import get_sample_request_by_request_no

REQUIRED_IMPORT_FIELDS = {"request_no", "tracking_number", "shipping_status", "shipped_at"}
KEX_TRACKING_FIELD = "เลขนำส่งพัสดุ"
KEX_REQUEST_NO_FIELD = "รหัสผู้รับ"
KEX_STATUS_FIELD = "สถานะการขนส่งสุดท้าย"
KEX_SHIPPED_AT_FIELD = "วันที่รับสินค้า"
KEX_FALLBACK_DATE_FIELD = "วันที่ยืนยันการสั่งซื้อ"


def normalize_header(value: object) -> str:
    return str(value or "").strip()


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def parse_optional_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.fromisoformat(value)


def map_kex_shipping_status(value: str) -> str:
    if "จัดส่งสำเร็จ" in value or "นำจ่ายสำเร็จ" in value:
        return ShippingStatus.DELIVERED.value
    if "ไม่สำเร็จ" in value or "ตีกลับ" in value:
        return ShippingStatus.FAILED.value
    if value:
        return ShippingStatus.SHIPPED.value
    return ShippingStatus.SHIPPED.value


def build_failed_job(
    db: Session,
    *,
    filename: str,
    created_by_user_id: int,
    error_message: str,
) -> ImportJob:
    job = ImportJob(
        filename=filename,
        total_rows=0,
        success_count=0,
        failed_count=1,
        not_found_count=0,
        created_by_user_id=created_by_user_id,
    )
    job.rows.append(
        ImportJobRow(
            row_number=0,
            request_no=None,
            tracking_number=None,
            status="failed",
            error_message=error_message,
        ),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def parse_csv_rows(content: bytes) -> tuple[list[dict[str, str]], str | None]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    missing_fields = REQUIRED_IMPORT_FIELDS.difference(reader.fieldnames or [])
    if missing_fields:
        return [], f"Missing required columns: {', '.join(sorted(missing_fields))}"
    return list(reader), None


def parse_kex_xlsx_rows(content: bytes) -> tuple[list[dict[str, str]], str | None]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [normalize_header(value) for value in header_row or []]
    required_headers = {KEX_TRACKING_FIELD, KEX_REQUEST_NO_FIELD, KEX_STATUS_FIELD}
    missing_headers = required_headers.difference(headers)
    if missing_headers:
        return [], f"Missing required columns: {', '.join(sorted(missing_headers))}"

    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        source = dict(zip(headers, values, strict=False))
        request_no = normalize_cell(source.get(KEX_REQUEST_NO_FIELD))
        tracking_number = normalize_cell(source.get(KEX_TRACKING_FIELD))
        shipping_status = map_kex_shipping_status(normalize_cell(source.get(KEX_STATUS_FIELD)))
        shipped_at = normalize_cell(source.get(KEX_SHIPPED_AT_FIELD))
        if not shipped_at:
            shipped_at = normalize_cell(source.get(KEX_FALLBACK_DATE_FIELD))
        if not request_no and not tracking_number:
            continue
        rows.append(
            {
                "request_no": request_no,
                "tracking_number": tracking_number,
                "shipping_status": shipping_status,
                "shipped_at": shipped_at,
            },
        )
    return rows, None


def parse_tracking_rows(filename: str, content: bytes) -> tuple[list[dict[str, str]], str | None]:
    if filename.lower().endswith(".xlsx"):
        return parse_kex_xlsx_rows(content)
    return parse_csv_rows(content)


def import_tracking_csv(
    db: Session,
    *,
    filename: str,
    content: bytes,
    created_by_user_id: int,
) -> ImportJob:
    rows, parse_error = parse_tracking_rows(filename, content)
    if parse_error:
        return build_failed_job(
            db,
            filename=filename,
            created_by_user_id=created_by_user_id,
            error_message=parse_error,
        )

    job = ImportJob(
        filename=filename,
        total_rows=0,
        success_count=0,
        failed_count=0,
        not_found_count=0,
        created_by_user_id=created_by_user_id,
    )
    db.add(job)

    for row_number, row in enumerate(rows, start=2):
        job.total_rows += 1
        request_no = (row.get("request_no") or "").strip()
        tracking_number = (row.get("tracking_number") or "").strip()
        shipping_status = (row.get("shipping_status") or "").strip()

        try:
            parsed_shipping_status = ShippingStatus(shipping_status)
            sample_request = get_sample_request_by_request_no(db, request_no)
            if sample_request is None:
                job.not_found_count += 1
                job.rows.append(
                    ImportJobRow(
                        row_number=row_number,
                        request_no=request_no,
                        tracking_number=tracking_number,
                        status="not_found",
                        error_message="Sample request not found.",
                    ),
                )
                continue

            sample_request.tracking_number = tracking_number
            sample_request.shipping_status = parsed_shipping_status.value
            if parsed_shipping_status == ShippingStatus.SHIPPED:
                sample_request.request_status = RequestStatus.SHIPPED.value
            sample_request.shipped_at = parse_optional_datetime(row.get("shipped_at"))
            sample_request.updated_by_user_id = created_by_user_id
            job.success_count += 1
            job.rows.append(
                ImportJobRow(
                    row_number=row_number,
                    request_no=request_no,
                    tracking_number=tracking_number,
                    status="success",
                    error_message=None,
                ),
            )
        except (ValueError, TypeError) as exc:
            job.failed_count += 1
            job.rows.append(
                ImportJobRow(
                    row_number=row_number,
                    request_no=request_no or None,
                    tracking_number=tracking_number or None,
                    status="failed",
                    error_message=str(exc),
                ),
            )

    db.commit()
    db.refresh(job)
    return job
